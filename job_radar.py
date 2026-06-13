#!/usr/bin/env python
"""
Personal Job Radar

Reads companies.xlsx + profile.json + resume.txt/resume.pdf, fetches company jobs,
scores them locally, and updates output/job_matches.xlsx.
"""

from __future__ import annotations

import argparse
import dataclasses
import datetime as dt
import html
import json
import os
import re
import sys
import time
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover - optional outside Codex bundled runtime
    PdfReader = None


ROOT = Path(__file__).resolve().parent
INPUT_COMPANIES = ROOT / "companies.xlsx"
PROFILE_PATH = ROOT / "profile.json"
RESUME_TXT = ROOT / "resume.txt"
RESUME_PDF = ROOT / "resume.pdf"
OUTPUT_DIR = ROOT / "output"
OUTPUT_XLSX = OUTPUT_DIR / "job_matches.xlsx"

STATUSES = ["Not Applied", "Applied", "Interview", "Rejected", "Offer", "Skip"]


DEFAULT_PROFILE = {
    "target_titles": [
        "backend engineer",
        "software engineer",
        "genai engineer",
        "ai platform engineer",
        "ml platform engineer",
        "platform engineer",
        "infrastructure engineer",
        "search engineer",
    ],
    "preferred_keywords": [
        "python",
        "java",
        "backend",
        "distributed systems",
        "microservices",
        "kubernetes",
        "aws",
        "gcp",
        "llm",
        "genai",
        "rag",
        "agents",
        "retrieval",
        "search",
        "platform",
        "infrastructure",
    ],
    "avoid_keywords": [
        "intern",
        "internship",
        "frontend only",
        "qa",
        "quality assurance",
        "support",
        "sales engineer",
        "customer success",
        "director",
        "principal",
    ],
    "preferred_locations": ["india", "bengaluru", "bangalore", "remote"],
    "max_experience_years": 6,
    "posted_within_days": 30,
    "minimum_score": 70,
}


SAMPLE_COMPANIES = [
    ["OpenAI", "https://jobs.ashbyhq.com/OpenAI", 1, "Verified Ashby example; replace with your target company", "yes"],
    ["Airbnb", "https://boards.greenhouse.io/airbnb", 2, "Verified Greenhouse example; replace with your target company", "yes"],
    ["Reddit", "https://boards.greenhouse.io/reddit", 2, "Verified Greenhouse example; replace with your target company", "yes"],
    ["Databricks", "https://boards.greenhouse.io/databricks", 2, "Verified Greenhouse example; replace with your target company", "no"],
]


@dataclasses.dataclass
class Job:
    company: str
    title: str
    location: str = ""
    department: str = ""
    job_url: str = ""
    apply_url: str = ""
    posted_date: str = ""
    description: str = ""
    employment_type: str = ""
    ats_type: str = "generic"
    scraped_at: str = ""
    job_id: str = ""
    score: int = 0
    reason: str = ""
    matched_skills: str = ""
    gaps: str = ""
    status: str = "Not Applied"
    notes: str = ""


def now_iso() -> str:
    return dt.datetime.now().replace(microsecond=0).isoformat(sep=" ")


def today() -> dt.date:
    return dt.datetime.now().date()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value)
    text = re.sub(r"<script[\s\S]*?</script>", " ", text, flags=re.I)
    text = re.sub(r"<style[\s\S]*?</style>", " ", text, flags=re.I)
    text = re.sub(r"<[^>]+>", " ", text)
    text = html.unescape(text)
    return re.sub(r"\s+", " ", text).strip()


def slug(text: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", text.lower()).strip("-")


def safe_sheet_name(name: str, used: set[str]) -> str:
    base = re.sub(r"[\[\]\:\*\?\/\\]", " ", name).strip() or "Company"
    base = base[:28]
    candidate = base
    i = 2
    while candidate in used:
        suffix = f" {i}"
        candidate = f"{base[:31 - len(suffix)]}{suffix}"
        i += 1
    used.add(candidate)
    return candidate


def read_json(path: Path, default: dict[str, Any]) -> dict[str, Any]:
    if not path.exists():
        path.write_text(json.dumps(default, indent=2), encoding="utf-8")
        return default
    with path.open("r", encoding="utf-8") as f:
        data = json.load(f)
    merged = default.copy()
    merged.update(data)
    return merged


def create_companies_template(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "companies"
    headers = ["company", "career_url", "priority", "notes", "enabled"]
    ws.append(headers)
    for row in SAMPLE_COMPANIES:
        ws.append(row)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    widths = [24, 72, 10, 28, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    wb.save(path)


def read_companies(path: Path) -> list[dict[str, Any]]:
    if not path.exists():
        create_companies_template(path)
    wb = load_workbook(path)
    ws = wb["companies"] if "companies" in wb.sheetnames else wb.active
    headers = [str(c.value).strip().lower() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        item = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        if not item.get("company") or not item.get("career_url"):
            continue
        enabled = str(item.get("enabled", "yes")).strip().lower()
        if enabled in {"no", "false", "0", "n"}:
            continue
        rows.append(item)
    return rows


def read_resume_text() -> str:
    if RESUME_TXT.exists():
        return RESUME_TXT.read_text(encoding="utf-8", errors="ignore")
    if RESUME_PDF.exists() and PdfReader is not None:
        reader = PdfReader(str(RESUME_PDF))
        return "\n".join(page.extract_text() or "" for page in reader.pages)
    if not RESUME_TXT.exists():
        RESUME_TXT.write_text(
            "Paste your resume text here, or place resume.pdf next to job_radar.py.\n",
            encoding="utf-8",
        )
    return ""


def http_json(url: str, timeout: int = 25) -> Any:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 PersonalJobRadar/1.0",
            "Accept": "application/json,text/html;q=0.9,*/*;q=0.8",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        body = response.read().decode("utf-8", errors="replace")
    return json.loads(body)


def http_text(url: str, timeout: int = 25) -> str:
    req = urllib.request.Request(
        url,
        headers={
            "User-Agent": "Mozilla/5.0 PersonalJobRadar/1.0",
            "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        },
    )
    with urllib.request.urlopen(req, timeout=timeout) as response:
        return response.read().decode("utf-8", errors="replace")


def url_path_parts(url: str) -> list[str]:
    parsed = urllib.parse.urlparse(url)
    return [p for p in parsed.path.split("/") if p]


def fetch_greenhouse(company: str, url: str) -> list[Job]:
    parts = url_path_parts(url)
    token = parts[0] if parts else ""
    if not token:
        return []
    api = f"https://boards-api.greenhouse.io/v1/boards/{token}/jobs?content=true"
    data = http_json(api)
    jobs = []
    for item in data.get("jobs", []):
        offices = item.get("offices") or []
        departments = item.get("departments") or []
        location = (item.get("location") or {}).get("name") or ", ".join(o.get("name", "") for o in offices)
        job = Job(
            company=company,
            title=item.get("title", ""),
            location=location,
            department=", ".join(d.get("name", "") for d in departments),
            job_url=item.get("absolute_url", ""),
            apply_url=item.get("absolute_url", ""),
            posted_date=item.get("first_published") or item.get("updated_at") or "",
            description=clean_text(item.get("content", "")),
            employment_type="",
            ats_type="greenhouse",
            job_id=f"greenhouse:{item.get('id', '')}",
        )
        jobs.append(job)
    return jobs


def fetch_lever(company: str, url: str) -> list[Job]:
    parts = url_path_parts(url)
    token = parts[0] if parts else ""
    if not token:
        return []
    api = f"https://api.lever.co/v0/postings/{token}?mode=json"
    data = http_json(api)
    jobs = []
    for item in data:
        categories = item.get("categories") or {}
        created_at = item.get("createdAt")
        posted = ""
        if isinstance(created_at, int):
            posted = dt.datetime.fromtimestamp(created_at / 1000).date().isoformat()
        lists_text = " ".join(clean_text((l or {}).get("content", "")) for l in item.get("lists", []))
        jobs.append(
            Job(
                company=company,
                title=item.get("text", ""),
                location=categories.get("location", ""),
                department=categories.get("department", ""),
                job_url=item.get("hostedUrl", ""),
                apply_url=item.get("applyUrl") or item.get("hostedUrl", ""),
                posted_date=posted,
                description=f"{clean_text(item.get('description', ''))} {lists_text}",
                employment_type=categories.get("commitment", ""),
                ats_type="lever",
                job_id=f"lever:{item.get('id', '')}",
            )
        )
    return jobs


def fetch_ashby(company: str, url: str) -> list[Job]:
    parts = url_path_parts(url)
    token = parts[0] if parts else ""
    if not token:
        return []
    token_candidates = []
    for candidate in [token, urllib.parse.unquote(token), token.lower(), token.title(), company, company.replace(" ", "")]:
        if candidate and candidate not in token_candidates:
            token_candidates.append(candidate)
    data = None
    last_error: Exception | None = None
    for candidate in token_candidates:
        api = f"https://api.ashbyhq.com/posting-api/job-board/{urllib.parse.quote(candidate)}?includeCompensation=true"
        try:
            data = http_json(api)
            break
        except urllib.error.HTTPError as exc:
            last_error = exc
            if exc.code != 404:
                raise
        except Exception as exc:
            last_error = exc
            raise
    if data is None:
        if last_error:
            raise last_error
        return []
    jobs = []
    for item in data.get("jobs", []):
        location = item.get("location")
        if isinstance(location, dict):
            location = location.get("name", "")
        jobs.append(
            Job(
                company=company,
                title=item.get("title", ""),
                location=location or "",
                department=item.get("department", ""),
                job_url=item.get("jobUrl", ""),
                apply_url=item.get("jobUrl", ""),
                posted_date=item.get("publishedDate") or "",
                description=clean_text(item.get("descriptionHtml") or item.get("descriptionPlain") or ""),
                employment_type=item.get("employmentType", ""),
                ats_type="ashby",
                job_id=f"ashby:{item.get('id', '')}",
            )
        )
    return jobs


def fetch_smartrecruiters(company: str, url: str) -> list[Job]:
    parsed = urllib.parse.urlparse(url)
    parts = url_path_parts(url)
    token = ""
    if "jobs.smartrecruiters.com" in parsed.netloc and parts:
        token = parts[0]
    query = urllib.parse.parse_qs(parsed.query)
    if not token and "company" in query:
        token = query["company"][0]
    if not token:
        return []
    api = f"https://api.smartrecruiters.com/v1/companies/{token}/postings?limit=100"
    data = http_json(api)
    jobs = []
    for item in data.get("content", []):
        location = item.get("location") or {}
        loc_text = ", ".join(str(location.get(k, "")) for k in ["city", "region", "country"] if location.get(k))
        job_url = item.get("ref") or item.get("applyUrl") or ""
        jobs.append(
            Job(
                company=company,
                title=item.get("name", ""),
                location=loc_text,
                department=item.get("department", {}).get("label", "") if isinstance(item.get("department"), dict) else "",
                job_url=job_url,
                apply_url=item.get("applyUrl") or job_url,
                posted_date=item.get("releasedDate") or "",
                description=clean_text(item.get("jobAd", {}).get("sections", {}).get("jobDescription", "")) if isinstance(item.get("jobAd"), dict) else "",
                employment_type=item.get("typeOfEmployment", {}).get("label", "") if isinstance(item.get("typeOfEmployment"), dict) else "",
                ats_type="smartrecruiters",
                job_id=f"smartrecruiters:{item.get('id', '')}",
            )
        )
    return jobs


def fetch_generic(company: str, url: str) -> list[Job]:
    text = http_text(url)
    jobs = []

    for block in re.findall(r'<script[^>]+type=["\']application/ld\+json["\'][^>]*>([\s\S]*?)</script>', text, re.I):
        try:
            data = json.loads(html.unescape(block.strip()))
        except Exception:
            continue
        candidates = data if isinstance(data, list) else [data]
        for item in candidates:
            if not isinstance(item, dict) or item.get("@type") != "JobPosting":
                continue
            loc = item.get("jobLocation", "")
            if isinstance(loc, list):
                loc = ", ".join(clean_text(x) for x in loc)
            elif isinstance(loc, dict):
                address = loc.get("address", {})
                if isinstance(address, dict):
                    loc = ", ".join(str(address.get(k, "")) for k in ["addressLocality", "addressRegion", "addressCountry"] if address.get(k))
                else:
                    loc = clean_text(address)
            jobs.append(
                Job(
                    company=company,
                    title=item.get("title", ""),
                    location=clean_text(loc),
                    department=item.get("industry", ""),
                    job_url=item.get("url") or url,
                    apply_url=item.get("url") or url,
                    posted_date=item.get("datePosted", ""),
                    description=clean_text(item.get("description", "")),
                    employment_type=item.get("employmentType", ""),
                    ats_type="generic-jsonld",
                    job_id=f"generic:{slug(company)}:{slug(item.get('title', ''))}:{slug(str(loc))}",
                )
            )

    if jobs:
        return jobs

    # Lightweight fallback: collect job-like links from the page.
    seen = set()
    for href, label in re.findall(r'<a[^>]+href=["\']([^"\']+)["\'][^>]*>([\s\S]*?)</a>', text, re.I):
        title = clean_text(label)
        if len(title) < 4 or len(title) > 120:
            continue
        blob = f"{title} {href}".lower()
        if not any(word in blob for word in ["engineer", "developer", "software", "backend", "platform", "machine learning", "genai", "ai"]):
            continue
        full_url = urllib.parse.urljoin(url, html.unescape(href))
        key = (title.lower(), full_url)
        if key in seen:
            continue
        seen.add(key)
        jobs.append(
            Job(
                company=company,
                title=title,
                job_url=full_url,
                apply_url=full_url,
                posted_date="",
                description=title,
                ats_type="generic-links",
                job_id=f"generic:{slug(company)}:{slug(title)}",
            )
        )
    return jobs


def detect_and_fetch(company: str, url: str) -> tuple[list[Job], str, str]:
    parsed = urllib.parse.urlparse(url)
    host = parsed.netloc.lower()
    lower_url = url.lower()
    try:
        if "greenhouse.io" in host:
            return fetch_greenhouse(company, url), "greenhouse", ""
        if "lever.co" in host:
            return fetch_lever(company, url), "lever", ""
        if "ashbyhq.com" in host:
            return fetch_ashby(company, url), "ashby", ""
        if "smartrecruiters.com" in host or "smartrecruiters" in lower_url:
            return fetch_smartrecruiters(company, url), "smartrecruiters", ""
        return fetch_generic(company, url), "generic", ""
    except urllib.error.HTTPError as exc:
        return [], "error", f"HTTP {exc.code}: {exc.reason}"
    except urllib.error.URLError as exc:
        return [], "error", f"Network error: {exc.reason}"
    except Exception as exc:
        return [], "error", f"{type(exc).__name__}: {exc}"


def parse_date(value: str) -> dt.date | None:
    if not value:
        return None
    value = str(value).strip()
    patterns = [
        r"(\d{4}-\d{2}-\d{2})",
        r"(\d{4}/\d{2}/\d{2})",
    ]
    for pattern in patterns:
        match = re.search(pattern, value)
        if match:
            raw = match.group(1).replace("/", "-")
            try:
                return dt.date.fromisoformat(raw)
            except ValueError:
                pass
    return None


def count_phrase_matches(phrases: list[str], text: str) -> list[str]:
    text_l = text.lower()
    found = []
    for phrase in phrases:
        p = phrase.lower().strip()
        if not p:
            continue
        if re.search(rf"(?<![a-z0-9]){re.escape(p)}(?![a-z0-9])", text_l):
            found.append(phrase)
    return found


def infer_experience_years(text: str) -> int | None:
    years = []
    for match in re.finditer(r"(\d{1,2})\+?\s*(?:years|yrs|year)\b", text.lower()):
        try:
            years.append(int(match.group(1)))
        except ValueError:
            pass
    return min(years) if years else None


def score_job(job: Job, profile: dict[str, Any], resume_text: str) -> None:
    searchable = f"{job.title} {job.location} {job.department} {job.employment_type} {job.description}".lower()
    resume_l = resume_text.lower()

    target_titles = profile.get("target_titles", [])
    preferred = profile.get("preferred_keywords", [])
    avoid = profile.get("avoid_keywords", [])
    locations = profile.get("preferred_locations", [])
    max_years = int(profile.get("max_experience_years", 99))
    posted_within_days = int(profile.get("posted_within_days", 3650))

    score = 0
    reasons = []
    gaps = []

    title_matches = count_phrase_matches(target_titles, job.title)
    if title_matches:
        score += 25
        reasons.append(f"title match: {', '.join(title_matches[:3])}")
    elif any(w in job.title.lower() for w in ["engineer", "developer", "software", "platform", "ai", "ml"]):
        score += 12
        reasons.append("engineering title")
    else:
        gaps.append("title not strongly aligned")

    skill_matches = count_phrase_matches(preferred, searchable)
    resume_skill_matches = [s for s in skill_matches if s.lower() in resume_l or not resume_text.strip()]
    skill_score = min(30, len(skill_matches) * 4 + len(resume_skill_matches) * 2)
    score += skill_score
    if skill_matches:
        reasons.append(f"skills: {', '.join(skill_matches[:8])}")
    else:
        gaps.append("few preferred skills found")

    location_matches = count_phrase_matches(locations, job.location)
    if not job.location:
        score += 5
        reasons.append("location unknown")
    elif location_matches:
        score += 15
        reasons.append(f"location: {', '.join(location_matches[:3])}")
    else:
        gaps.append(f"location may not match: {job.location}")

    posted = parse_date(job.posted_date)
    if posted:
        age = (today() - posted).days
        if age <= posted_within_days:
            score += 15
            reasons.append(f"recent: {age} days old")
        elif age <= posted_within_days * 2:
            score += 7
            reasons.append(f"older: {age} days old")
        else:
            gaps.append(f"old posting: {age} days")
    else:
        score += 5
        reasons.append("posted date unknown")

    exp = infer_experience_years(searchable)
    if exp is None:
        score += 8
    elif exp <= max_years:
        score += 10
        reasons.append(f"experience <= {max_years} years")
    else:
        score -= 15
        gaps.append(f"asks {exp}+ years")

    avoid_matches = count_phrase_matches(avoid, searchable)
    if avoid_matches:
        penalty = min(35, len(avoid_matches) * 12)
        score -= penalty
        gaps.append(f"avoid terms: {', '.join(avoid_matches[:5])}")

    if re.search(r"\b(full[- ]?time|permanent)\b", searchable):
        score += 5

    job.score = max(0, min(100, int(score)))
    job.reason = "; ".join(reasons[:5])
    job.matched_skills = ", ".join(skill_matches[:12])
    job.gaps = "; ".join(gaps[:5])


def stable_key(job: Job) -> str:
    if job.job_id and not job.job_id.endswith(":"):
        return job.job_id.lower()
    if job.apply_url:
        return job.apply_url.lower()
    return f"{job.company}|{job.title}|{job.location}".lower()


def read_previous_statuses(path: Path) -> dict[str, dict[str, str]]:
    if not path.exists():
        return {}
    try:
        wb = load_workbook(path, data_only=False)
    except Exception:
        return {}
    statuses = {}
    for sheet_name in ["All Jobs", "Top Matches", "Applied Tracker"]:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
        idx = {h: i for i, h in enumerate(headers)}
        required = {"Job Key", "Status", "Notes"}
        if not required.issubset(idx):
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = row[idx["Job Key"]]
            if not key:
                continue
            statuses[str(key).lower()] = {
                "Status": str(row[idx["Status"]] or "Not Applied"),
                "Notes": str(row[idx["Notes"]] or ""),
            }
    return statuses


def row_for(job: Job) -> list[Any]:
    return [
        job.score,
        job.company,
        job.title,
        job.location,
        job.posted_date,
        job.apply_url or job.job_url,
        job.reason,
        job.matched_skills,
        job.gaps,
        job.status,
        job.notes,
        job.ats_type,
        job.department,
        job.employment_type,
        job.scraped_at,
        stable_key(job),
    ]


HEADERS = [
    "Score",
    "Company",
    "Title",
    "Location",
    "Posted Date",
    "Apply Link",
    "Reason",
    "Matched Skills",
    "Gaps",
    "Status",
    "Notes",
    "ATS",
    "Department",
    "Employment Type",
    "Last Checked",
    "Job Key",
]


def style_sheet(ws, max_row: int, hyperlink_col: int = 6) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = [10, 22, 42, 24, 16, 34, 52, 34, 38, 16, 30, 16, 22, 18, 22, 34]
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "Z"].width = width
    for row in ws.iter_rows(min_row=2, max_row=max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        link_cell = row[hyperlink_col - 1]
        if link_cell.value:
            link_cell.hyperlink = str(link_cell.value)
            link_cell.style = "Hyperlink"
    if max_row >= 2:
        score_range = f"A2:A{max_row}"
        ws.conditional_formatting.add(score_range, CellIsRule(operator="greaterThanOrEqual", formula=["90"], fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["80", "89"], fill=PatternFill("solid", fgColor="D9EAD3")))
        ws.conditional_formatting.add(score_range, CellIsRule(operator="between", formula=["70", "79"], fill=PatternFill("solid", fgColor="FFF2CC")))
        ws.conditional_formatting.add(score_range, CellIsRule(operator="lessThan", formula=["70"], fill=PatternFill("solid", fgColor="F4CCCC")))
        dv = DataValidation(type="list", formula1=f'"{",".join(STATUSES)}"', allow_blank=False)
        ws.add_data_validation(dv)
        dv.add(f"J2:J{max_row}")


def write_jobs_sheet(wb: Workbook, name: str, jobs: list[Job]) -> None:
    ws = wb.create_sheet(name)
    ws.append(HEADERS)
    for job in jobs:
        ws.append(row_for(job))
    style_sheet(ws, max(1, ws.max_row))


def write_company_status(wb: Workbook, statuses: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet("Company Status")
    headers = ["Company", "Career URL", "ATS", "Jobs Found", "Last Checked", "Error"]
    ws.append(headers)
    for item in statuses:
        ws.append([item.get(h, "") for h in headers])
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col, width in zip("ABCDEF", [24, 72, 18, 12, 22, 60]):
        ws.column_dimensions[col].width = width
    for row in ws.iter_rows(min_row=2):
        row[1].hyperlink = row[1].value
        row[1].style = "Hyperlink"
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def write_output(jobs: list[Job], company_statuses: list[dict[str, Any]], profile: dict[str, Any]) -> None:
    OUTPUT_DIR.mkdir(exist_ok=True)
    minimum_score = int(profile.get("minimum_score", 70))
    top = [j for j in jobs if j.score >= minimum_score]
    low = [j for j in jobs if j.score < minimum_score]
    applied = [j for j in jobs if j.status and j.status != "Not Applied"]

    wb = Workbook()
    wb.remove(wb.active)
    write_jobs_sheet(wb, "Top Matches", sorted(top, key=lambda j: (-j.score, j.company, j.title)))
    write_jobs_sheet(wb, "All Jobs", sorted(jobs, key=lambda j: (-j.score, j.company, j.title)))
    write_jobs_sheet(wb, "Low Match", sorted(low, key=lambda j: (-j.score, j.company, j.title)))
    write_jobs_sheet(wb, "Applied Tracker", sorted(applied, key=lambda j: (j.company, j.title)))

    used = set(wb.sheetnames)
    for company in sorted({j.company for j in jobs}):
        company_jobs = [j for j in jobs if j.company == company]
        write_jobs_sheet(wb, safe_sheet_name(company, used), sorted(company_jobs, key=lambda j: -j.score))

    write_company_status(wb, company_statuses)
    tmp = OUTPUT_XLSX.with_suffix(".tmp.xlsx")
    wb.save(tmp)
    if OUTPUT_XLSX.exists():
        OUTPUT_XLSX.unlink()
    tmp.rename(OUTPUT_XLSX)


def scan() -> int:
    profile = read_json(PROFILE_PATH, DEFAULT_PROFILE)
    companies = read_companies(INPUT_COMPANIES)
    resume_text = read_resume_text()
    previous = read_previous_statuses(OUTPUT_XLSX)

    all_jobs: list[Job] = []
    company_statuses: list[dict[str, Any]] = []
    checked_at = now_iso()

    for company_item in companies:
        company = str(company_item.get("company", "")).strip()
        url = str(company_item.get("career_url", "")).strip()
        jobs, ats, error = detect_and_fetch(company, url)
        for job in jobs:
            job.scraped_at = checked_at
            key = stable_key(job)
            if key in previous:
                job.status = previous[key].get("Status") or "Not Applied"
                job.notes = previous[key].get("Notes") or ""
            score_job(job, profile, resume_text)
        all_jobs.extend(jobs)
        company_statuses.append(
            {
                "Company": company,
                "Career URL": url,
                "ATS": ats,
                "Jobs Found": len(jobs),
                "Last Checked": checked_at,
                "Error": error,
            }
        )
        print(f"{company}: {len(jobs)} jobs ({ats})" + (f" - {error}" if error else ""))
        time.sleep(0.4)

    write_output(all_jobs, company_statuses, profile)
    print(f"\nUpdated: {OUTPUT_XLSX}")
    print(f"Jobs scanned: {len(all_jobs)}")
    print(f"Top matches: {sum(1 for j in all_jobs if j.score >= int(profile.get('minimum_score', 70)))}")
    return 0


def main() -> int:
    parser = argparse.ArgumentParser(description="Update Excel job matches from company career pages.")
    parser.add_argument("--init", action="store_true", help="Create companies.xlsx/profile.json/resume.txt templates and exit.")
    args = parser.parse_args()

    read_json(PROFILE_PATH, DEFAULT_PROFILE)
    if not INPUT_COMPANIES.exists():
        create_companies_template(INPUT_COMPANIES)
    if not RESUME_TXT.exists() and not RESUME_PDF.exists():
        RESUME_TXT.write_text(
            "Paste your resume text here, or place resume.pdf in this folder.\n",
            encoding="utf-8",
        )

    if args.init:
        print(f"Ready. Edit {INPUT_COMPANIES}, {PROFILE_PATH}, and {RESUME_TXT} or add resume.pdf.")
        return 0

    return scan()


if __name__ == "__main__":
    raise SystemExit(main())
